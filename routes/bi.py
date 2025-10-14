from flask import Blueprint, render_template, request, send_file, make_response
from flask_login import login_required, current_user
from models import db
from models.empresa import Empresa
from models.visita import Visita
from models.demanda import Demanda
from models.inovacao import InovacaoEmpresa
from models.diagnostico import Diagnostico
from models.user import User
from sqlalchemy import func, extract
from datetime import datetime, timedelta
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import openpyxl

bi_bp = Blueprint('bi', __name__, url_prefix='/bi')

@bi_bp.route('/')
@login_required
def index():
    periodo = request.args.get('periodo', '30')
    
    data_inicio = datetime.now() - timedelta(days=int(periodo))
    
    total_empresas = Empresa.query.count()
    empresas_ativas = Empresa.query.filter_by(status='Ativo').count()
    
    visitas_periodo = Visita.query.filter(Visita.data_visita >= data_inicio.date()).count()
    demandas_periodo = Demanda.query.filter(Demanda.data_abertura >= data_inicio.date()).count()
    
    inovacoes_periodo = InovacaoEmpresa.query.filter(
        InovacaoEmpresa.data_inicio >= data_inicio.date()
    ).count()
    
    diagnosticos_periodo = Diagnostico.query.filter(
        Diagnostico.data_diagnostico >= data_inicio.date()
    ).count()
    
    economia_total = db.session.query(
        func.sum(InovacaoEmpresa.economia_gerada)
    ).scalar() or 0
    
    investimento_total = db.session.query(
        func.sum(InovacaoEmpresa.investimento)
    ).scalar() or 0
    
    roi_geral = ((economia_total - investimento_total) / investimento_total * 100) if investimento_total > 0 else 0
    
    empresas_por_porte = db.session.query(
        Empresa.porte,
        func.count(Empresa.id)
    ).group_by(Empresa.porte).all()
    
    empresas_por_uf = db.session.query(
        Empresa.uf,
        func.count(Empresa.id)
    ).group_by(Empresa.uf).order_by(func.count(Empresa.id).desc()).limit(10).all()
    
    demandas_por_tipo = db.session.query(
        Demanda.tipo,
        func.count(Demanda.id)
    ).group_by(Demanda.tipo).all()
    
    demandas_por_status = db.session.query(
        Demanda.status,
        func.count(Demanda.id)
    ).group_by(Demanda.status).all()
    
    evolucao_mensal = db.session.query(
        extract('month', Visita.data_visita).label('mes'),
        func.count(Visita.id).label('total')
    ).filter(
        Visita.data_visita >= (datetime.now() - timedelta(days=365)).date()
    ).group_by('mes').order_by('mes').all()
    
    consultores_ativos = User.query.filter_by(perfil='Consultor', ativo=True).count()
    
    top_consultores = db.session.query(
        User.nome,
        func.count(Diagnostico.id).label('total')
    ).join(Diagnostico, User.id == Diagnostico.consultor_id).group_by(
        User.id, User.nome
    ).order_by(func.count(Diagnostico.id).desc()).limit(10).all()
    
    return render_template('bi/index.html',
                         total_empresas=total_empresas,
                         empresas_ativas=empresas_ativas,
                         visitas_periodo=visitas_periodo,
                         demandas_periodo=demandas_periodo,
                         inovacoes_periodo=inovacoes_periodo,
                         diagnosticos_periodo=diagnosticos_periodo,
                         economia_total=economia_total,
                         investimento_total=investimento_total,
                         roi_geral=roi_geral,
                         empresas_por_porte=empresas_por_porte,
                         empresas_por_uf=empresas_por_uf,
                         demandas_por_tipo=demandas_por_tipo,
                         demandas_por_status=demandas_por_status,
                         evolucao_mensal=evolucao_mensal,
                         consultores_ativos=consultores_ativos,
                         top_consultores=top_consultores,
                         periodo=periodo)

@bi_bp.route('/exportar/excel')
@login_required
def exportar_excel():
    wb = openpyxl.Workbook()
    
    ws_empresas = wb.active
    ws_empresas.title = 'Empresas'
    ws_empresas.append(['Nome', 'CNPJ', 'Porte', 'Cidade', 'UF', 'Status'])
    
    empresas = Empresa.query.all()
    for emp in empresas:
        ws_empresas.append([emp.nome, emp.cnpj, emp.porte, emp.cidade, emp.uf, emp.status])
    
    ws_demandas = wb.create_sheet('Demandas')
    ws_demandas.append(['Empresa', 'Titulo', 'Tipo', 'Status', 'Prioridade', 'Data Abertura'])
    
    demandas = Demanda.query.join(Empresa).all()
    for dem in demandas:
        ws_demandas.append([
            dem.empresa.nome,
            dem.titulo,
            dem.tipo,
            dem.status,
            dem.prioridade,
            str(dem.data_abertura)
        ])
    
    ws_inovacoes = wb.create_sheet('Inovacoes')
    ws_inovacoes.append(['Empresa', 'Inovacao', 'Status', 'Investimento', 'Economia', 'ROI %'])
    
    inovacoes = InovacaoEmpresa.query.join(Empresa).all()
    for inov in inovacoes:
        roi = ((inov.economia_gerada - inov.investimento) / inov.investimento * 100) if inov.investimento > 0 else 0
        ws_inovacoes.append([
            inov.empresa.nome,
            inov.inovacao.nome,
            inov.status,
            inov.investimento,
            inov.economia_gerada,
            round(roi, 2)
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'relatorio_completo_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@bi_bp.route('/exportar/pdf')
@login_required
def exportar_pdf():
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    titulo = Paragraph('<b>Relatorio Gerencial - SRI SENAI</b>', styles['Title'])
    elements.append(titulo)
    elements.append(Spacer(1, 20))
    
    data_atual = Paragraph(f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal'])
    elements.append(data_atual)
    elements.append(Spacer(1, 20))
    
    empresas_data = [['Nome', 'CNPJ', 'Cidade', 'Status']]
    empresas = Empresa.query.limit(50).all()
    for emp in empresas:
        empresas_data.append([
            emp.nome[:30],
            emp.cnpj,
            emp.cidade[:20],
            emp.status
        ])
    
    tabela = Table(empresas_data)
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(tabela)
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'relatorio_{datetime.now().strftime("%Y%m%d")}.pdf'
    )
